"""
Claude Monitoring System
A professional dashboard for monitoring Claude Memory System v2.0
"""

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, Response, send_file
from flask_socketio import SocketIO, emit
from functools import wraps
import os
import csv
import io
import bcrypt
import threading
import time
from datetime import datetime
from utils.metrics import MetricsCollector
from utils.log_parser import LogParser
from utils.policy_checker import PolicyChecker
from utils.session_tracker import SessionTracker
from utils.history_tracker import HistoryTracker
from utils.notification_manager import NotificationManager
from utils.alert_sender import AlertSender
from utils.community_widgets import CommunityWidgetsManager
from utils.anomaly_detector import AnomalyDetector
from utils.predictive_analytics import PredictiveAnalytics
from utils.alert_routing import AlertRoutingEngine
from utils.memory_system_monitor import MemorySystemMonitor
from utils.widget_version_manager import WidgetVersionManager
from utils.widget_comments_manager import WidgetCommentsManager
from utils.collaboration_manager import CollaborationSessionManager
from utils.trending_calculator import TrendingCalculator
from flasgger import Swagger, swag_from
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

app = Flask(__name__)
app.secret_key = 'claude-monitoring-system-secret-key-2026'

# Initialize SocketIO for real-time updates
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Initialize Swagger for API documentation
swagger_config = {
    "headers": [],
    "specs": [
        {
            "endpoint": 'apispec',
            "route": '/api/docs/apispec.json',
            "rule_filter": lambda rule: True,
            "model_filter": lambda tag: True,
        }
    ],
    "static_url_path": "/flasgger_static",
    "swagger_ui": True,
    "specs_route": "/api/docs"
}
swagger = Swagger(app, config=swagger_config)

# Initialize utilities
metrics = MetricsCollector()
log_parser = LogParser()
policy_checker = PolicyChecker()
session_tracker = SessionTracker()
history_tracker = HistoryTracker()
notification_manager = NotificationManager()
alert_sender = AlertSender()
community_widgets_manager = CommunityWidgetsManager()
anomaly_detector = AnomalyDetector()
predictive_analytics = PredictiveAnalytics()
alert_routing = AlertRoutingEngine()
memory_system_monitor = MemorySystemMonitor()
widget_version_manager = WidgetVersionManager()
widget_comments_manager = WidgetCommentsManager()
collaboration_manager = CollaborationSessionManager()
trending_calculator = TrendingCalculator()

# User database (in production, use a proper database)
# Password: 'admin' (hashed with bcrypt)
USERS = {
    'admin': {
        'password_hash': bcrypt.hashpw('admin'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
        'role': 'admin'
    }
}

def verify_password(username, password):
    """Verify username and password"""
    if username not in USERS:
        return False
    stored_hash = USERS[username]['password_hash'].encode('utf-8')
    return bcrypt.checkpw(password.encode('utf-8'), stored_hash)

def update_password(username, new_password):
    """Update user password"""
    if username not in USERS:
        return False
    USERS[username]['password_hash'] = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    return True

# ============================================================
# Analytics Helper Functions
# ============================================================

def calculate_trend(values):
    """Calculate trend percentage change"""
    if not values or len(values) < 2:
        return {'direction': 'stable', 'percentage': 0}

    first_half = values[:len(values)//2]
    second_half = values[len(values)//2:]

    first_avg = sum(first_half) / len(first_half) if first_half else 0
    second_avg = sum(second_half) / len(second_half) if second_half else 0

    if first_avg == 0:
        return {'direction': 'stable', 'percentage': 0}

    change = ((second_avg - first_avg) / first_avg) * 100

    direction = 'up' if change > 5 else ('down' if change < -5 else 'stable')

    return {
        'direction': direction,
        'percentage': round(abs(change), 1),
        'current': round(second_avg, 1),
        'previous': round(first_avg, 1)
    }

def calculate_policy_effectiveness():
    """Calculate policy effectiveness metrics"""
    try:
        optimization_impact = metrics.get_optimization_impact()
        total_opts = optimization_impact.get('total_optimizations', 0)

        if total_opts == 0:
            return {'effectiveness': 0, 'total_interventions': 0}

        return {
            'effectiveness': min(100, (total_opts / 100) * 100),  # Normalize to percentage
            'total_interventions': total_opts,
            'context_optimizations': optimization_impact.get('context_optimizations', 0),
            'failures_prevented': optimization_impact.get('failures_prevented', 0),
            'model_selections': optimization_impact.get('model_selections', 0)
        }
    except:
        return {'effectiveness': 0, 'total_interventions': 0}

def calculate_daemon_uptime(daemon_status):
    """Calculate daemon uptime percentage"""
    if not daemon_status:
        return 0

    running = len([d for d in daemon_status if d.get('status') == 'running'])
    total = len(daemon_status)

    return round((running / total) * 100, 1) if total > 0 else 0

def calculate_peak_hours(historical_data):
    """Calculate peak usage hours (placeholder)"""
    # This is a simplified version - in production, you'd analyze actual usage patterns
    return {
        'peak_hour': '10:00 AM - 11:00 AM',
        'peak_day': 'Monday',
        'busiest_period': 'Morning (9 AM - 12 PM)'
    }

def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    """Redirect to dashboard or login"""
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Login page
    ---
    tags:
      - Authentication
    parameters:
      - name: username
        in: formData
        type: string
        required: true
        description: Username
      - name: password
        in: formData
        type: string
        required: true
        description: Password
    responses:
      200:
        description: Login page or redirect to dashboard
      302:
        description: Redirect to dashboard on successful login
    """
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if verify_password(username, password):
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('dashboard'))
        else:
            error = 'Invalid credentials. Please try again.'

    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    """Logout"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard"""
    # Get time range from query parameter (default: 7 days)
    days = request.args.get('days', 7, type=int)
    # Validate days parameter
    if days not in [7, 30, 60, 90]:
        days = 7

    # Get widget preferences from session (default: all enabled)
    widget_prefs = session.get('widget_preferences', {
        'system_health': True,
        'daemon_status': True,
        'policy_status': True,
        'recent_activity': True,
        'historical_charts': True,
        'recent_errors': True
    })

    # Get all metrics
    system_health = metrics.get_system_health()
    daemon_status = metrics.get_daemon_status()
    policy_status = policy_checker.get_all_policies_status()
    recent_activity = log_parser.get_recent_activity(limit=10)

    # Add daily metric snapshot
    try:
        # Get detailed policy status (returns dict)
        policy_status_dict = policy_checker.get_detailed_policy_status()

        current_metrics = {
            'health_score': system_health.get('health_score', 0),
            'errors_24h': log_parser.get_error_count(hours=24),
            'policy_hits': policy_status_dict.get('active_policies', 0),
            'context_usage': system_health.get('context_usage', 0),
            'tokens_used': 0,  # Would come from session tracker
            'daemons_running': len([d for d in daemon_status if d.get('status') == 'running']),
            'daemons_total': len(daemon_status)
        }
        history_tracker.add_daily_metric(current_metrics)
    except Exception as e:
        print(f"Error adding daily metric: {e}")
        import traceback
        traceback.print_exc()

    # Get historical data for selected time range
    chart_data = history_tracker.get_chart_data(days=days)
    summary_stats = history_tracker.get_summary_stats(days=days)

    return render_template('dashboard.html',
                         system_health=system_health,
                         daemon_status=daemon_status,
                         policy_status=policy_status,
                         recent_activity=recent_activity,
                         chart_data=chart_data,
                         summary_stats=summary_stats,
                         selected_days=days,
                         widget_preferences=widget_prefs)

@app.route('/comparison')
@login_required
def comparison():
    """Before/After comparison page"""
    comparison_data = metrics.get_cost_comparison()
    optimization_impact = metrics.get_optimization_impact()

    return render_template('comparison.html',
                         comparison=comparison_data,
                         impact=optimization_impact)

@app.route('/policies')
@login_required
def policies():
    """Policies status page"""
    policies_data = policy_checker.get_detailed_policy_status()
    policy_history = log_parser.get_policy_history()

    return render_template('policies.html',
                         policies=policies_data,
                         history=policy_history)

@app.route('/logs')
@login_required
def logs():
    """Log analyzer page"""
    log_files = log_parser.get_available_logs()

    return render_template('logs.html',
                         log_files=log_files)

@app.route('/api/logs/analyze', methods=['POST'])
@login_required
def analyze_logs():
    """API endpoint to analyze logs"""
    data = request.get_json()
    log_file = data.get('log_file')
    search_term = data.get('search_term', '')
    log_level = data.get('log_level', 'all')

    results = log_parser.analyze_log_file(log_file, search_term, log_level)

    return jsonify(results)

@app.route('/api/metrics')
@login_required
def api_metrics():
    """API endpoint for dashboard metrics - REAL DATA from Claude Memory System"""
    try:
        system_health = metrics.get_system_health()
        daemon_status = metrics.get_daemon_status()
        # Use get_detailed_policy_status() which returns a dict
        policy_status = policy_checker.get_detailed_policy_status()

        # daemon_status is a list from get_daemon_status
        daemons_running = len([d for d in daemon_status if isinstance(d, dict) and d.get('status') == 'running'])
        daemons_total = len(daemon_status) if daemon_status else 8

        health_score = system_health.get('health_score', system_health.get('score', 0))

        return jsonify({
            'success': True,
            'health_score': health_score,
            'daemons_running': daemons_running,
            'daemons_total': daemons_total,
            'active_policies': policy_status.get('active_policies', 0),
            'policy_hits': 0,  # Will track from logs later
            'context_usage': system_health.get('context_usage', 0),
            'memory_usage': system_health.get('memory_usage', 0),
            'labels': ['Now'],
            'health_scores': [health_score],
            'policy_hits_data': [0]
        })
    except Exception as e:
        print(f"Error in api_metrics: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/activity')
@login_required
def api_activity():
    """API endpoint for recent activity"""
    try:
        recent_activity = log_parser.get_recent_activity(limit=10)
        return jsonify({
            'success': True,
            'activities': recent_activity
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/policies')
@login_required
def api_policies():
    """API endpoint for policy status - REAL DATA"""
    try:
        policies_data = policy_checker.get_detailed_policy_status()

        # Ensure policies is an array for dashboard
        if isinstance(policies_data, dict):
            # Convert dict to array of policy objects
            policies_array = []
            for key, value in policies_data.items():
                if isinstance(value, dict):
                    policies_array.append({
                        'name': key.replace('_', ' ').title(),
                        **value
                    })
                else:
                    # Simple key-value pair
                    policies_array.append({
                        'name': key.replace('_', ' ').title(),
                        'status': 'active',
                        'value': value
                    })
            policies_data = policies_array

        return jsonify({
            'success': True,
            'policies': policies_data
        })
    except Exception as e:
        print(f"Error in api_policies: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/system-info')
@login_required
def api_system_info():
    """API endpoint for system information - REAL DATA"""
    try:
        system_health = metrics.get_system_health()
        daemon_status = metrics.get_daemon_status()

        daemons_running = len([d for d in daemon_status if d.get('status') == 'running'])
        daemons_total = len(daemon_status)
        health_score = system_health.get('health_score', system_health.get('score', 0))

        return jsonify({
            'success': True,
            'system_info': {
                'status': 'Operational' if health_score >= 90 else ('Healthy' if health_score >= 70 else 'Degraded'),
                'health_score': health_score,
                'memory_usage': system_health.get('memory_usage', 0),
                'context_usage': system_health.get('context_usage', 0),
                'daemons_running': daemons_running,
                'daemons_total': daemons_total,
                'uptime': system_health.get('uptime', 'Active'),
                'last_check': datetime.now().isoformat()
            }
        })
    except Exception as e:
        print(f"Error in api_system_info: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/recent-errors')
@login_required
def api_recent_errors():
    """API endpoint for recent errors"""
    try:
        errors = log_parser.get_recent_errors(limit=5)
        return jsonify({
            'success': True,
            'errors': errors
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/comparison')
@login_required
def api_comparison():
    """API endpoint for comparison data"""
    try:
        comparison_data = metrics.get_cost_comparison()
        optimization_impact = metrics.get_optimization_impact()

        return jsonify({
            'success': True,
            'comparison': comparison_data,
            'impact': optimization_impact
        })
    except Exception as e:
        print(f"Error in api_comparison: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/metrics/live')
@login_required
def live_metrics():
    """API endpoint for live metrics"""
    live_data = {
        'health_score': metrics.get_health_score(),
        'daemon_count': metrics.get_running_daemon_count(),
        'context_usage': metrics.get_context_usage(),
        'recent_errors': log_parser.get_error_count(hours=1),
        'timestamp': datetime.now().isoformat()
    }

    return jsonify(live_data)

@app.route('/api/daemon/restart/<daemon_name>', methods=['POST'])
@login_required
def restart_daemon(daemon_name):
    """API endpoint to restart daemon"""
    try:
        result = metrics.restart_daemon(daemon_name)
        return jsonify({'success': True, 'message': f'Daemon {daemon_name} restarted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/sessions')
@login_required
def sessions():
    """Sessions tracking page"""
    current_session = session_tracker.update_session_metrics()
    sessions_history = session_tracker.get_sessions_history()
    last_session = session_tracker.get_last_session()

    # Compare current with last
    comparison = None
    if last_session:
        comparison = session_tracker.compare_sessions(last_session, current_session)

    summary = session_tracker.get_all_sessions_summary()

    return render_template('sessions.html',
                         current_session=current_session,
                         last_session=last_session,
                         sessions_history=sessions_history[-10:],  # Last 10 sessions
                         comparison=comparison,
                         summary=summary)

@app.route('/api/session/end', methods=['POST'])
@login_required
def end_session():
    """API endpoint to end current session"""
    try:
        ended_session = session_tracker.end_current_session()
        return jsonify({'success': True, 'session': ended_session})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/settings')
@login_required
def settings():
    """
    Settings page
    ---
    tags:
      - Settings
    responses:
      200:
        description: Settings page
    """
    # Get alert thresholds from session
    alert_thresholds = session.get('alert_thresholds', {
        'health_score': 70,
        'error_count': 10,
        'context_usage': 85,
        'daemon_down': True
    })

    # Get current theme
    current_theme = session.get('dashboard_theme', 'default')

    return render_template('settings.html',
                         alert_thresholds=alert_thresholds,
                         current_theme=current_theme)

@app.route('/api/themes', methods=['GET', 'POST'])
@login_required
def dashboard_themes():
    """
    Get or set dashboard theme
    ---
    tags:
      - Themes
    parameters:
      - name: body
        in: body
        required: false
        schema:
          type: object
          properties:
            theme:
              type: string
              description: Theme name (default/dark/blue/purple/green/orange)
    responses:
      200:
        description: Theme saved or retrieved
    """
    if request.method == 'POST':
        try:
            data = request.get_json()
            theme = data.get('theme', 'default')
            session['dashboard_theme'] = theme
            return jsonify({'success': True, 'message': 'Theme saved', 'theme': theme})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    else:
        theme = session.get('dashboard_theme', 'default')
        return jsonify({'success': True, 'theme': theme})

@app.route('/widgets')
@login_required
def widgets():
    """
    Widget Marketplace
    ---
    tags:
      - Widgets
    responses:
      200:
        description: Widget marketplace page
    """
    return render_template('widgets.html')

@app.route('/widget-builder')
@login_required
def widget_builder():
    """
    Advanced Widget Builder
    ---
    tags:
      - Widgets
    responses:
      200:
        description: Widget builder page with visual editor
    """
    return render_template('widget-builder.html')

@app.route('/api/widgets/save', methods=['POST'])
@login_required
def save_custom_widget():
    """
    Save custom widget from builder
    ---
    tags:
      - Widgets
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            name:
              type: string
              description: Widget name
            components:
              type: array
              description: Widget components
            config:
              type: object
              description: Widget configuration
            html:
              type: string
              description: Widget HTML
            css:
              type: string
              description: Widget CSS
            js:
              type: string
              description: Widget JavaScript
    responses:
      200:
        description: Widget saved successfully
    """
    try:
        data = request.get_json()

        required_fields = ['name', 'html']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field} is required'}), 400

        # Get custom widgets from session
        custom_widgets = session.get('custom_widgets_advanced', [])

        # Generate widget ID
        widget_id = f"advanced-{len(custom_widgets) + 1}"

        widget = {
            'id': widget_id,
            'name': data['name'],
            'components': data.get('components', []),
            'config': data.get('config', {}),
            'html': data['html'],
            'css': data.get('css', ''),
            'js': data.get('js', ''),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }

        custom_widgets.append(widget)
        session['custom_widgets_advanced'] = custom_widgets

        return jsonify({
            'success': True,
            'message': 'Widget saved successfully',
            'widget': widget
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/community-marketplace')
@login_required
def community_marketplace():
    """
    Community Widget Marketplace
    ---
    tags:
      - Community
    responses:
      200:
        description: Community marketplace page
    """
    return render_template('community-marketplace.html')

@app.route('/api/community-widgets', methods=['GET'])
@login_required
def get_community_widgets():
    """
    Get all community widgets
    ---
    tags:
      - Community
    responses:
      200:
        description: List of community widgets
    """
    try:
        widgets = community_widgets_manager.get_all_widgets()
        return jsonify({
            'success': True,
            'widgets': widgets,
            'count': len(widgets)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/community-widgets/stats', methods=['GET'])
@login_required
def get_community_stats():
    """
    Get community marketplace statistics
    ---
    tags:
      - Community
    responses:
      200:
        description: Marketplace statistics
    """
    try:
        stats = community_widgets_manager.get_stats()
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/community-widgets/publish', methods=['POST'])
@login_required
def publish_widget():
    """
    Publish widget to community
    ---
    tags:
      - Community
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            name:
              type: string
              description: Widget name
            description:
              type: string
              description: Widget description
            category:
              type: string
              description: Widget category
            version:
              type: string
              description: Widget version
            tags:
              type: array
              description: Widget tags
            author:
              type: string
              description: Author name
            widget_data:
              type: object
              description: Widget data/definition
    responses:
      200:
        description: Widget published successfully
    """
    try:
        data = request.get_json()

        required_fields = ['name', 'description', 'category', 'widget_data']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field} is required'}), 400

        widget = community_widgets_manager.publish_widget(data)

        return jsonify({
            'success': True,
            'message': 'Widget published to community successfully',
            'widget': widget
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/community-widgets/<widget_id>/download', methods=['GET'])
@login_required
def download_community_widget(widget_id):
    """
    Download a community widget
    ---
    tags:
      - Community
    parameters:
      - name: widget_id
        in: path
        type: string
        required: true
        description: Widget ID
    responses:
      200:
        description: Widget data for download
    """
    try:
        widget = community_widgets_manager.get_widget_by_id(widget_id)

        if not widget:
            return jsonify({'success': False, 'message': 'Widget not found'}), 404

        # Increment download count
        community_widgets_manager.increment_downloads(widget_id)

        return jsonify({
            'success': True,
            'name': widget['name'],
            'widget_data': widget['widget_data']
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/community-widgets/<widget_id>/rate', methods=['POST'])
@login_required
def rate_community_widget(widget_id):
    """
    Rate a community widget
    ---
    tags:
      - Community
    parameters:
      - name: widget_id
        in: path
        type: string
        required: true
        description: Widget ID
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            rating:
              type: integer
              description: Rating value (1-5)
    responses:
      200:
        description: Rating submitted successfully
    """
    try:
        data = request.get_json()
        rating = data.get('rating')

        if not rating or not (1 <= rating <= 5):
            return jsonify({'success': False, 'message': 'Rating must be between 1 and 5'}), 400

        success = community_widgets_manager.add_rating(widget_id, rating)

        if not success:
            return jsonify({'success': False, 'message': 'Widget not found'}), 404

        return jsonify({
            'success': True,
            'message': 'Rating submitted successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Widget Version Control API
# ============================================================

@app.route('/api/widgets/<widget_id>/versions/create', methods=['POST'])
@login_required
def create_widget_version(widget_id):
    """Create a new version of a widget"""
    try:
        data = request.get_json()
        widget_data = data.get('widget_data')
        version_type = data.get('version_type', 'patch')
        commit_message = data.get('commit_message', '')

        username = session.get('username', 'admin')

        version = widget_version_manager.create_version(
            widget_id=widget_id,
            widget_data=widget_data,
            version_type=version_type,
            commit_message=commit_message,
            created_by=username
        )

        return jsonify({
            'success': True,
            'version': version
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/versions', methods=['GET'])
@login_required
def get_widget_versions(widget_id):
    """Get all versions for a widget"""
    try:
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))

        versions = widget_version_manager.get_version_list(widget_id, limit, offset)

        return jsonify({
            'success': True,
            'versions': versions,
            'total': len(versions)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/versions/<version>', methods=['GET'])
@login_required
def get_widget_version(widget_id, version):
    """Get specific version data"""
    try:
        version_data = widget_version_manager.get_version(widget_id, version)

        if not version_data:
            return jsonify({'success': False, 'message': 'Version not found'}), 404

        return jsonify({
            'success': True,
            'version_data': version_data
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/versions/<version>/diff', methods=['GET'])
@login_required
def get_widget_version_diff(widget_id, version):
    """Get diff between current and specified version"""
    try:
        from_version = request.args.get('from', version)
        to_version = request.args.get('to', widget_version_manager.get_current_version(widget_id).get('version'))

        diff = widget_version_manager.get_diff(widget_id, from_version, to_version)

        return jsonify({
            'success': True,
            'diff': diff
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/versions/<version>/rollback', methods=['POST'])
@login_required
def rollback_widget_version(widget_id, version):
    """Rollback widget to a previous version"""
    try:
        username = session.get('username', 'admin')

        new_version = widget_version_manager.rollback_version(
            widget_id=widget_id,
            target_version=version,
            created_by=username
        )

        return jsonify({
            'success': True,
            'message': f'Rolled back to version {version}',
            'new_version': new_version
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/versions/<version>', methods=['DELETE'])
@login_required
def delete_widget_version(widget_id, version):
    """Delete a specific version"""
    try:
        success = widget_version_manager.delete_version(widget_id, version)

        if not success:
            return jsonify({'success': False, 'message': 'Version not found or cannot be deleted'}), 404

        return jsonify({
            'success': True,
            'message': f'Version {version} deleted successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Widget Comments & Discussions API
# ============================================================

@app.route('/api/widgets/<widget_id>/comments', methods=['POST'])
@login_required
def add_widget_comment(widget_id):
    """Add a comment to a widget"""
    try:
        data = request.get_json()
        content = data.get('content')
        parent_comment_id = data.get('parent_comment_id')

        if not content or len(content.strip()) == 0:
            return jsonify({'success': False, 'message': 'Comment cannot be empty'}), 400

        username = session.get('username', 'admin')

        comment = widget_comments_manager.add_comment(
            widget_id=widget_id,
            author=username,
            content=content,
            parent_comment_id=parent_comment_id
        )

        # Emit real-time notification
        socketio.emit('comment:new', {
            'widget_id': widget_id,
            'comment': comment
        }, broadcast=True)

        return jsonify({
            'success': True,
            'comment': comment
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/comments', methods=['GET'])
@login_required
def get_widget_comments(widget_id):
    """Get all comments for a widget"""
    try:
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        thread_id = request.args.get('thread_id')

        comments = widget_comments_manager.get_comments(
            widget_id=widget_id,
            limit=limit,
            offset=offset,
            thread_id=thread_id
        )

        return jsonify({
            'success': True,
            'comments': comments,
            'total': len(comments)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/comments/<comment_id>', methods=['PUT'])
@login_required
def update_widget_comment(widget_id, comment_id):
    """Update a comment"""
    try:
        data = request.get_json()
        new_content = data.get('content')

        if not new_content or len(new_content.strip()) == 0:
            return jsonify({'success': False, 'message': 'Comment cannot be empty'}), 400

        username = session.get('username', 'admin')

        comment = widget_comments_manager.update_comment(
            widget_id=widget_id,
            comment_id=comment_id,
            author=username,
            new_content=new_content
        )

        if not comment:
            return jsonify({'success': False, 'message': 'Comment not found or unauthorized'}), 404

        return jsonify({
            'success': True,
            'comment': comment
        })
    except PermissionError as e:
        return jsonify({'success': False, 'message': str(e)}), 403
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/comments/<comment_id>', methods=['DELETE'])
@login_required
def delete_widget_comment(widget_id, comment_id):
    """Delete a comment"""
    try:
        username = session.get('username', 'admin')
        is_admin = USERS.get(username, {}).get('role') == 'admin'

        success = widget_comments_manager.delete_comment(
            widget_id=widget_id,
            comment_id=comment_id,
            author=username,
            is_admin=is_admin
        )

        if not success:
            return jsonify({'success': False, 'message': 'Comment not found'}), 404

        return jsonify({
            'success': True,
            'message': 'Comment deleted successfully'
        })
    except PermissionError as e:
        return jsonify({'success': False, 'message': str(e)}), 403
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/comments/<comment_id>/react', methods=['POST'])
@login_required
def add_comment_reaction(widget_id, comment_id):
    """Add a reaction to a comment"""
    try:
        data = request.get_json()
        reaction_type = data.get('reaction_type', 'thumbs_up')

        username = session.get('username', 'admin')

        comment = widget_comments_manager.add_reaction(
            widget_id=widget_id,
            comment_id=comment_id,
            user=username,
            reaction_type=reaction_type
        )

        if not comment:
            return jsonify({'success': False, 'message': 'Comment not found'}), 404

        return jsonify({
            'success': True,
            'comment': comment
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/notifications/comments', methods=['GET'])
@login_required
def get_comment_notifications():
    """Get comment notifications for current user"""
    try:
        username = session.get('username', 'admin')
        limit = int(request.args.get('limit', 20))

        mentions = widget_comments_manager.get_user_mentions(username, limit)

        return jsonify({
            'success': True,
            'mentions': mentions,
            'total': len(mentions)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Real-time Collaboration API
# ============================================================

@app.route('/api/widgets/<widget_id>/collaborate/start', methods=['POST'])
@login_required
def start_collaboration_session(widget_id):
    """Start a new collaboration session"""
    try:
        data = request.get_json()
        duration_hours = data.get('duration_hours', 2)

        username = session.get('username', 'admin')

        session_data = collaboration_manager.create_session(
            widget_id=widget_id,
            creator=username,
            session_duration_hours=duration_hours
        )

        return jsonify({
            'success': True,
            'session': session_data
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/collaborate/<session_id>/join', methods=['POST'])
@login_required
def join_collaboration_session(widget_id, session_id):
    """Join an existing collaboration session"""
    try:
        data = request.get_json()
        socket_id = data.get('socket_id', '')

        username = session.get('username', 'admin')

        session_data = collaboration_manager.join_session(
            session_id=session_id,
            user_id=username,
            socket_id=socket_id
        )

        if not session_data:
            return jsonify({'success': False, 'message': 'Session not found or expired'}), 404

        return jsonify({
            'success': True,
            'session': session_data
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/collaborate/<session_id>/leave', methods=['POST'])
@login_required
def leave_collaboration_session(widget_id, session_id):
    """Leave a collaboration session"""
    try:
        username = session.get('username', 'admin')

        success = collaboration_manager.leave_session(
            session_id=session_id,
            user_id=username
        )

        if not success:
            return jsonify({'success': False, 'message': 'Session not found'}), 404

        return jsonify({
            'success': True,
            'message': 'Left session successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/collaborate/<session_id>/status', methods=['GET'])
@login_required
def get_collaboration_status(widget_id, session_id):
    """Get collaboration session status"""
    try:
        session_data = collaboration_manager.get_session(session_id)

        if not session_data:
            return jsonify({'success': False, 'message': 'Session not found'}), 404

        return jsonify({
            'success': True,
            'session': session_data
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Featured & Trending Widgets API
# ============================================================

@app.route('/api/widgets/featured', methods=['GET'])
@login_required
def get_featured_widgets():
    """Get featured widgets"""
    try:
        featured = trending_calculator.get_featured_widgets()

        return jsonify({
            'success': True,
            'featured': featured,
            'total': len(featured)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/trending', methods=['GET'])
@login_required
def get_trending_widgets():
    """Get trending widgets"""
    try:
        period_days = int(request.args.get('period', 1))
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'

        trending = trending_calculator.get_trending_cached(
            time_period_days=period_days,
            force_refresh=force_refresh
        )

        return jsonify({
            'success': True,
            'trending': trending,
            'period_days': period_days,
            'total': len(trending)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/feature', methods=['POST'])
@login_required
def feature_widget(widget_id):
    """Feature a widget (admin only)"""
    try:
        username = session.get('username', 'admin')
        is_admin = USERS.get(username, {}).get('role') == 'admin'

        if not is_admin:
            return jsonify({'success': False, 'message': 'Admin access required'}), 403

        success = trending_calculator.add_featured(widget_id, username)

        if not success:
            return jsonify({'success': False, 'message': 'Widget not found or already featured'}), 404

        return jsonify({
            'success': True,
            'message': 'Widget featured successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/<widget_id>/feature', methods=['DELETE'])
@login_required
def unfeature_widget(widget_id):
    """Remove featured status (admin only)"""
    try:
        username = session.get('username', 'admin')
        is_admin = USERS.get(username, {}).get('role') == 'admin'

        if not is_admin:
            return jsonify({'success': False, 'message': 'Admin access required'}), 403

        success = trending_calculator.remove_featured(widget_id)

        if not success:
            return jsonify({'success': False, 'message': 'Widget not found in featured list'}), 404

        return jsonify({
            'success': True,
            'message': 'Featured status removed'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/trending/calculate', methods=['POST'])
@login_required
def recalculate_trending():
    """Recalculate trending widgets (admin only)"""
    try:
        username = session.get('username', 'admin')
        is_admin = USERS.get(username, {}).get('role') == 'admin'

        if not is_admin:
            return jsonify({'success': False, 'message': 'Admin access required'}), 403

        trending_calculator.invalidate_cache()

        # Recalculate for all periods
        trending_24h = trending_calculator.get_trending_cached(1, force_refresh=True)
        trending_7d = trending_calculator.get_trending_cached(7, force_refresh=True)
        trending_30d = trending_calculator.get_trending_cached(30, force_refresh=True)

        return jsonify({
            'success': True,
            'message': 'Trending data recalculated',
            'stats': {
                'trending_24h': len(trending_24h),
                'trending_7d': len(trending_7d),
                'trending_30d': len(trending_30d)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/install', methods=['POST'])
@login_required
def install_widget():
    """
    Install a widget
    ---
    tags:
      - Widgets
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            widget_id:
              type: string
              description: Widget ID to install
    responses:
      200:
        description: Widget installed successfully
    """
    try:
        data = request.get_json()
        widget_id = data.get('widget_id')

        if not widget_id:
            return jsonify({'success': False, 'message': 'Widget ID is required'}), 400

        # Get installed widgets from session
        installed_widgets = session.get('installed_widgets', [])

        # Check if already installed
        if widget_id in installed_widgets:
            return jsonify({'success': False, 'message': 'Widget already installed'})

        # Add to installed list
        installed_widgets.append(widget_id)
        session['installed_widgets'] = installed_widgets

        return jsonify({
            'success': True,
            'message': 'Widget installed successfully',
            'widget_id': widget_id
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/installed')
@login_required
def get_installed_widgets():
    """
    Get installed widgets
    ---
    tags:
      - Widgets
    responses:
      200:
        description: List of installed widgets
    """
    installed_widgets = session.get('installed_widgets', [])

    # Widget metadata (simplified)
    widget_metadata = {
        'health-score-meter': {'name': 'Health Score Meter', 'category': 'metrics'},
        'error-trends-chart': {'name': 'Error Trends', 'category': 'charts'},
        'cost-tracker': {'name': 'Cost Tracker', 'category': 'metrics'},
        'policy-monitor': {'name': 'Policy Monitor', 'category': 'metrics'},
        'alert-feed': {'name': 'Live Alert Feed', 'category': 'alerts'},
        'context-monitor': {'name': 'Context Monitor', 'category': 'tools'},
        'session-timeline': {'name': 'Session Timeline', 'category': 'tools'},
        'model-distribution': {'name': 'Model Distribution', 'category': 'charts'},
        'quick-actions': {'name': 'Quick Actions', 'category': 'tools'}
    }

    widgets = []
    for widget_id in installed_widgets:
        if widget_id in widget_metadata:
            widget = widget_metadata[widget_id].copy()
            widget['id'] = widget_id
            widgets.append(widget)

    return jsonify({
        'success': True,
        'widgets': widgets,
        'count': len(widgets)
    })

@app.route('/api/widgets/create', methods=['POST'])
@login_required
def create_custom_widget():
    """
    Create a custom widget
    ---
    tags:
      - Widgets
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            name:
              type: string
              description: Widget name
            description:
              type: string
              description: Widget description
            category:
              type: string
              description: Widget category
            icon:
              type: string
              description: Font Awesome icon class
            color:
              type: string
              description: Widget color
            data_source:
              type: string
              description: API endpoint for data
    responses:
      200:
        description: Custom widget created successfully
    """
    try:
        data = request.get_json()

        required_fields = ['name', 'description', 'category', 'icon', 'color']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field} is required'}), 400

        # Get custom widgets from session
        custom_widgets = session.get('custom_widgets', [])

        # Generate widget ID
        widget_id = f"custom-{len(custom_widgets) + 1}"

        widget = {
            'id': widget_id,
            'name': data['name'],
            'description': data['description'],
            'category': data['category'],
            'icon': data['icon'],
            'color': data['color'],
            'data_source': data.get('data_source', ''),
            'created_at': datetime.now().isoformat()
        }

        custom_widgets.append(widget)
        session['custom_widgets'] = custom_widgets

        # Auto-install the custom widget
        installed_widgets = session.get('installed_widgets', [])
        installed_widgets.append(widget_id)
        session['installed_widgets'] = installed_widgets

        return jsonify({
            'success': True,
            'message': 'Custom widget created successfully',
            'widget': widget
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widgets/uninstall', methods=['POST'])
@login_required
def uninstall_widget():
    """
    Uninstall a widget
    ---
    tags:
      - Widgets
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            widget_id:
              type: string
              description: Widget ID to uninstall
    responses:
      200:
        description: Widget uninstalled successfully
    """
    try:
        data = request.get_json()
        widget_id = data.get('widget_id')

        if not widget_id:
            return jsonify({'success': False, 'message': 'Widget ID is required'}), 400

        # Get installed widgets from session
        installed_widgets = session.get('installed_widgets', [])

        if widget_id not in installed_widgets:
            return jsonify({'success': False, 'message': 'Widget not installed'})

        # Remove from installed list
        installed_widgets.remove(widget_id)
        session['installed_widgets'] = installed_widgets

        return jsonify({
            'success': True,
            'message': 'Widget uninstalled successfully',
            'widget_id': widget_id
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/analytics')
@login_required
def analytics():
    """
    Advanced Analytics Dashboard
    ---
    tags:
      - Analytics
    responses:
      200:
        description: Analytics dashboard page
    """
    # Get time range
    days = request.args.get('days', 30, type=int)
    if days not in [7, 30, 60, 90]:
        days = 30

    # Get comprehensive analytics data
    system_health = metrics.get_system_health()
    daemon_status = metrics.get_daemon_status()

    # Historical data
    historical_data = history_tracker.get_last_n_days(days)
    chart_data = history_tracker.get_chart_data(days)
    summary_stats = history_tracker.get_summary_stats(days)

    # Calculate trends
    analytics_data = {
        'health_trend': calculate_trend(chart_data.get('health_scores', [])),
        'error_trend': calculate_trend(chart_data.get('errors', [])),
        'context_trend': calculate_trend(chart_data.get('context_usage', [])),
        'policy_effectiveness': calculate_policy_effectiveness(),
        'daemon_uptime': calculate_daemon_uptime(daemon_status),
        'peak_hours': calculate_peak_hours(historical_data),
        'cost_analysis': metrics.get_cost_comparison(),
        'optimization_impact': metrics.get_optimization_impact()
    }

    return render_template('analytics.html',
                         selected_days=days,
                         system_health=system_health,
                         daemon_status=daemon_status,
                         chart_data=chart_data,
                         summary_stats=summary_stats,
                         analytics_data=analytics_data)

@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    """
    Change user password
    ---
    tags:
      - Authentication
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            current_password:
              type: string
              description: Current password
            new_password:
              type: string
              description: New password
            confirm_password:
              type: string
              description: Confirm new password
    responses:
      200:
        description: Password changed successfully
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
      400:
        description: Invalid request
      401:
        description: Current password incorrect
    """
    try:
        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')
        username = session.get('username')

        # Validate input
        if not all([current_password, new_password, confirm_password]):
            return jsonify({'success': False, 'message': 'All fields are required'}), 400

        if new_password != confirm_password:
            return jsonify({'success': False, 'message': 'New passwords do not match'}), 400

        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400

        # Verify current password
        if not verify_password(username, current_password):
            return jsonify({'success': False, 'message': 'Current password is incorrect'}), 401

        # Update password
        if update_password(username, new_password):
            return jsonify({'success': True, 'message': 'Password changed successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to update password'}), 500

    except Exception as e:
        print(f"Error changing password: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export/sessions')
@login_required
def export_sessions():
    """Export session history to CSV"""
    try:
        sessions_history = session_tracker.get_sessions_history()

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        writer.writerow(['Session ID', 'Start Time', 'End Time', 'Duration (min)', 'Commands', 'Tokens Used', 'Cost ($)', 'Status'])

        # Write data
        for sess in sessions_history:
            writer.writerow([
                sess.get('session_id', 'N/A'),
                sess.get('start_time', 'N/A'),
                sess.get('end_time', 'N/A'),
                sess.get('duration_minutes', 0),
                sess.get('commands_executed', 0),
                sess.get('tokens_used', 0),
                sess.get('estimated_cost', 0),
                sess.get('status', 'N/A')
            ])

        # Create response
        response = Response(output.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = f'attachment; filename=claude_sessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        return response
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export/metrics')
@login_required
def export_metrics():
    """Export current metrics to CSV"""
    try:
        system_health = metrics.get_system_health()
        daemon_status = metrics.get_daemon_status()

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        writer.writerow(['Metric', 'Value', 'Status', 'Timestamp'])

        # Write system health data
        writer.writerow(['Health Score', system_health.get('health_score', 0), 'N/A', datetime.now().isoformat()])
        writer.writerow(['Memory Usage', system_health.get('memory_usage', 0), 'N/A', datetime.now().isoformat()])
        writer.writerow(['Active Daemons', len([d for d in daemon_status if d.get('status') == 'running']), 'N/A', datetime.now().isoformat()])
        writer.writerow(['Total Daemons', len(daemon_status), 'N/A', datetime.now().isoformat()])

        # Write daemon status
        writer.writerow([])
        writer.writerow(['Daemon Name', 'Status', 'PID', 'Uptime'])
        for daemon in daemon_status:
            writer.writerow([
                daemon.get('name', 'N/A'),
                daemon.get('status', 'N/A'),
                daemon.get('pid', 'N/A'),
                daemon.get('uptime', 'N/A')
            ])

        # Create response
        response = Response(output.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = f'attachment; filename=claude_metrics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        return response
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export/logs')
@login_required
def export_logs():
    """Export logs to CSV"""
    try:
        recent_activity = log_parser.get_recent_activity(limit=1000)

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        writer.writerow(['Timestamp', 'Level', 'Policy', 'Action', 'Message'])

        # Write log data
        for activity in recent_activity:
            writer.writerow([
                activity.get('timestamp', 'N/A'),
                activity.get('level', 'N/A'),
                activity.get('policy', 'N/A'),
                activity.get('action', 'N/A'),
                activity.get('message', 'N/A')
            ])

        # Create response
        response = Response(output.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = f'attachment; filename=claude_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        return response
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export/excel/<type>')
@login_required
def export_excel(type):
    """
    Export data to Excel format
    ---
    tags:
      - Export
    parameters:
      - name: type
        in: path
        type: string
        required: true
        description: Type of data to export (sessions/metrics/logs/analytics)
    responses:
      200:
        description: Excel file download
    """
    try:
        wb = Workbook()
        ws = wb.active

        # Styling
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        if type == 'sessions':
            ws.title = "Sessions"
            sessions_history = session_tracker.get_sessions_history()

            # Headers
            headers = ['Session ID', 'Start Time', 'End Time', 'Duration (min)', 'Commands', 'Tokens Used', 'Cost ($)', 'Status']
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Data
            for sess in sessions_history:
                ws.append([
                    sess.get('session_id', 'N/A'),
                    sess.get('start_time', 'N/A'),
                    sess.get('end_time', 'N/A'),
                    sess.get('duration_minutes', 0),
                    sess.get('commands_executed', 0),
                    sess.get('tokens_used', 0),
                    sess.get('estimated_cost', 0),
                    sess.get('status', 'N/A')
                ])

        elif type == 'metrics':
            ws.title = "Metrics"
            system_health = metrics.get_system_health()
            daemon_status = metrics.get_daemon_status()

            # Headers
            headers = ['Metric', 'Value', 'Status', 'Timestamp']
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Data
            ws.append(['Health Score', system_health.get('health_score', 0), 'N/A', datetime.now().isoformat()])
            ws.append(['Memory Usage', system_health.get('memory_usage', 0), 'N/A', datetime.now().isoformat()])
            ws.append(['Active Daemons', len([d for d in daemon_status if d.get('status') == 'running']), 'N/A', datetime.now().isoformat()])
            ws.append(['Total Daemons', len(daemon_status), 'N/A', datetime.now().isoformat()])

        elif type == 'logs':
            ws.title = "Logs"
            recent_activity = log_parser.get_recent_activity(limit=1000)

            # Headers
            headers = ['Timestamp', 'Level', 'Policy', 'Action', 'Message']
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Data
            for activity in recent_activity:
                ws.append([
                    activity.get('timestamp', 'N/A'),
                    activity.get('level', 'N/A'),
                    activity.get('policy', 'N/A'),
                    activity.get('action', 'N/A'),
                    activity.get('message', 'N/A')
                ])

        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f'claude_{type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error exporting Excel: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export/pdf/<type>')
@login_required
def export_pdf(type):
    """
    Export data to PDF format
    ---
    tags:
      - Export
    parameters:
      - name: type
        in: path
        type: string
        required: true
        description: Type of data to export (sessions/metrics/logs/analytics)
    responses:
      200:
        description: PDF file download
    """
    try:
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=1  # Center
        )

        # Title
        title = Paragraph(f'Claude Monitoring System - {type.title()} Report', title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))

        # Date
        date_text = Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal'])
        elements.append(date_text)
        elements.append(Spacer(1, 0.3*inch))

        if type == 'sessions':
            sessions_history = session_tracker.get_sessions_history()

            # Table data
            data = [['Session ID', 'Start Time', 'Duration (min)', 'Commands', 'Tokens', 'Cost ($)', 'Status']]

            for sess in sessions_history[-20:]:  # Last 20 sessions
                data.append([
                    sess.get('session_id', 'N/A')[:8],
                    sess.get('start_time', 'N/A')[:16],
                    str(sess.get('duration_minutes', 0)),
                    str(sess.get('commands_executed', 0)),
                    str(sess.get('tokens_used', 0)),
                    str(sess.get('estimated_cost', 0)),
                    sess.get('status', 'N/A')
                ])

        elif type == 'metrics':
            system_health = metrics.get_system_health()
            daemon_status = metrics.get_daemon_status()

            # Table data
            data = [['Metric', 'Value', 'Timestamp']]
            data.append(['Health Score', f"{system_health.get('health_score', 0)}%", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            data.append(['Memory Usage', f"{system_health.get('memory_usage', 0)}%", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            data.append(['Active Daemons', f"{len([d for d in daemon_status if d.get('status') == 'running'])}/{len(daemon_status)}", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        elif type == 'logs':
            recent_activity = log_parser.get_recent_activity(limit=50)

            # Table data
            data = [['Timestamp', 'Level', 'Policy', 'Message']]

            for activity in recent_activity:
                data.append([
                    activity.get('timestamp', 'N/A')[:16],
                    activity.get('level', 'N/A'),
                    activity.get('policy', 'N/A')[:15],
                    activity.get('message', 'N/A')[:40]
                ])

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        pdf_buffer.seek(0)

        filename = f'claude_{type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error exporting PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/widget-preferences', methods=['GET', 'POST'])
@login_required
def widget_preferences():
    """Get or update widget preferences"""
    if request.method == 'POST':
        try:
            data = request.get_json()
            session['widget_preferences'] = data.get('preferences', {})
            return jsonify({'success': True, 'message': 'Preferences saved'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    else:
        # GET request - return current preferences
        prefs = session.get('widget_preferences', {
            'system_health': True,
            'daemon_status': True,
            'policy_status': True,
            'recent_activity': True,
            'historical_charts': True,
            'recent_errors': True
        })
        return jsonify({'success': True, 'preferences': prefs})

@app.route('/api/alert-thresholds', methods=['GET', 'POST'])
@login_required
def alert_thresholds():
    """
    Get or update alert thresholds
    ---
    tags:
      - Alerts
    parameters:
      - name: body
        in: body
        required: false
        schema:
          type: object
          properties:
            health_score:
              type: integer
              description: Health score threshold (0-100)
            error_count:
              type: integer
              description: Error count threshold per hour
            context_usage:
              type: integer
              description: Context usage threshold percentage
            daemon_down:
              type: boolean
              description: Alert when daemon is down
    responses:
      200:
        description: Alert thresholds saved or retrieved
    """
    if request.method == 'POST':
        try:
            data = request.get_json()
            thresholds = {
                'health_score': data.get('health_score', 70),
                'error_count': data.get('error_count', 10),
                'context_usage': data.get('context_usage', 85),
                'daemon_down': data.get('daemon_down', True)
            }
            session['alert_thresholds'] = thresholds
            return jsonify({'success': True, 'message': 'Alert thresholds saved', 'thresholds': thresholds})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    else:
        # GET request - return current thresholds
        thresholds = session.get('alert_thresholds', {
            'health_score': 70,
            'error_count': 10,
            'context_usage': 85,
            'daemon_down': True
        })
        return jsonify({'success': True, 'thresholds': thresholds})

@app.route('/api/check-alerts')
@login_required
def check_alerts():
    """
    Check current metrics against alert thresholds
    ---
    tags:
      - Alerts
    responses:
      200:
        description: Alert status
        schema:
          type: object
          properties:
            alerts:
              type: array
              items:
                type: object
    """
    try:
        thresholds = session.get('alert_thresholds', {
            'health_score': 70,
            'error_count': 10,
            'context_usage': 85,
            'daemon_down': True
        })

        system_health = metrics.get_system_health()
        daemon_status = metrics.get_daemon_status()

        alerts = []

        # Check health score
        health_score = system_health.get('health_score', 0)
        if health_score < thresholds.get('health_score', 70):
            alerts.append({
                'type': 'health_score',
                'severity': 'warning' if health_score >= 50 else 'critical',
                'message': f'Health score is {health_score}% (threshold: {thresholds.get("health_score")}%)',
                'value': health_score,
                'threshold': thresholds.get('health_score')
            })

        # Check context usage
        context_usage = system_health.get('context_usage', 0)
        if context_usage > thresholds.get('context_usage', 85):
            alerts.append({
                'type': 'context_usage',
                'severity': 'warning',
                'message': f'Context usage is {context_usage}% (threshold: {thresholds.get("context_usage")}%)',
                'value': context_usage,
                'threshold': thresholds.get('context_usage')
            })

        # Check daemons
        if thresholds.get('daemon_down', True):
            down_daemons = [d for d in daemon_status if d.get('status') != 'running']
            if down_daemons:
                alerts.append({
                    'type': 'daemon_down',
                    'severity': 'critical',
                    'message': f'{len(down_daemons)} daemon(s) are down',
                    'daemons': [d.get('name') for d in down_daemons]
                })

        # Check error count (last hour)
        error_count = log_parser.get_error_count(hours=1)
        if error_count > thresholds.get('error_count', 10):
            alerts.append({
                'type': 'error_count',
                'severity': 'warning',
                'message': f'{error_count} errors in last hour (threshold: {thresholds.get("error_count")})',
                'value': error_count,
                'threshold': thresholds.get('error_count')
            })

        # Create notifications for new alerts
        for alert in alerts:
            # Add browser notification
            notification_manager.add_notification(
                notification_type=alert['type'],
                title=f"{alert['severity'].upper()}: {alert['type'].replace('_', ' ').title()}",
                message=alert['message'],
                severity=alert['severity'],
                data=alert
            )

            # Send email/SMS alert
            try:
                alert_sender.send_alert(
                    alert_type=alert['type'],
                    severity=alert['severity'],
                    title=alert['type'].replace('_', ' ').title(),
                    message=alert['message']
                )
            except Exception as e:
                print(f"Error sending email/SMS alert: {e}")

        return jsonify({
            'success': True,
            'alert_count': len(alerts),
            'alerts': alerts,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        print(f"Error checking alerts: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/notifications')
@login_required
def notifications():
    """
    Notifications page
    ---
    tags:
      - Notifications
    responses:
      200:
        description: Notifications page
    """
    recent_notifications = notification_manager.get_recent_notifications(limit=50)
    unread_count = notification_manager.get_unread_count()
    trends = notification_manager.get_notification_trends(days=30)

    return render_template('notifications.html',
                         notifications=recent_notifications,
                         unread_count=unread_count,
                         trends=trends)

@app.route('/api/notifications')
@login_required
def api_notifications():
    """
    Get notifications
    ---
    tags:
      - Notifications
    parameters:
      - name: limit
        in: query
        type: integer
        default: 50
        description: Number of notifications to return
    responses:
      200:
        description: Notifications list
    """
    try:
        limit = request.args.get('limit', 50, type=int)
        notifications = notification_manager.get_recent_notifications(limit=limit)
        unread_count = notification_manager.get_unread_count()

        return jsonify({
            'success': True,
            'notifications': notifications,
            'unread_count': unread_count
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/notifications/<notification_id>/read', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    """
    Mark notification as read
    ---
    tags:
      - Notifications
    parameters:
      - name: notification_id
        in: path
        type: string
        required: true
    responses:
      200:
        description: Notification marked as read
    """
    try:
        notification_manager.mark_as_read(notification_id)
        return jsonify({'success': True, 'message': 'Notification marked as read'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    """
    Mark all notifications as read
    ---
    tags:
      - Notifications
    responses:
      200:
        description: All notifications marked as read
    """
    try:
        notification_manager.mark_all_as_read()
        return jsonify({'success': True, 'message': 'All notifications marked as read'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/notification-trends')
@login_required
def api_notification_trends():
    """
    Get notification trends
    ---
    tags:
      - Notifications
    parameters:
      - name: days
        in: query
        type: integer
        default: 30
        description: Number of days to analyze
    responses:
      200:
        description: Notification trends
    """
    try:
        days = request.args.get('days', 30, type=int)
        trends = notification_manager.get_notification_trends(days=days)

        return jsonify({
            'success': True,
            'trends': trends
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-config', methods=['GET', 'POST'])
@login_required
def alert_config():
    """
    Get or update alert configuration
    ---
    tags:
      - Alerts
    parameters:
      - name: body
        in: body
        required: false
        schema:
          type: object
          properties:
            email:
              type: object
              description: Email configuration
            sms:
              type: object
              description: SMS configuration
            alert_rules:
              type: object
              description: Alert rules configuration
    responses:
      200:
        description: Alert configuration retrieved or updated
    """
    if request.method == 'POST':
        try:
            data = request.get_json()
            config = alert_sender.load_config()

            # Update config with provided data
            if 'email' in data:
                config['email'] = {**config.get('email', {}), **data['email']}
            if 'sms' in data:
                config['sms'] = {**config.get('sms', {}), **data['sms']}
            if 'alert_rules' in data:
                config['alert_rules'] = {**config.get('alert_rules', {}), **data['alert_rules']}

            alert_sender.save_config(config)

            return jsonify({
                'success': True,
                'message': 'Alert configuration saved',
                'config': config
            })
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    else:
        try:
            config = alert_sender.load_config()
            # Remove sensitive data
            safe_config = config.copy()
            if 'email' in safe_config and 'password' in safe_config['email']:
                safe_config['email']['password'] = '***HIDDEN***' if safe_config['email']['password'] else ''
            if 'sms' in safe_config and 'auth_token' in safe_config['sms']:
                safe_config['sms']['auth_token'] = '***HIDDEN***' if safe_config['sms']['auth_token'] else ''

            return jsonify({
                'success': True,
                'config': safe_config
            })
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/test-email', methods=['POST'])
@login_required
def test_email():
    """
    Send test email
    ---
    tags:
      - Alerts
    responses:
      200:
        description: Test email sent
    """
    try:
        config = alert_sender.load_config()
        result = alert_sender.test_email(config)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/test-sms', methods=['POST'])
@login_required
def test_sms():
    """
    Send test SMS
    ---
    tags:
      - Alerts
    responses:
      200:
        description: Test SMS sent
    """
    try:
        config = alert_sender.load_config()
        result = alert_sender.test_sms(config)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/send-alert', methods=['POST'])
@login_required
def send_alert_manual():
    """
    Manually send an alert
    ---
    tags:
      - Alerts
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            alert_type:
              type: string
              description: Alert type
            severity:
              type: string
              description: Severity (critical, warning, info)
            title:
              type: string
              description: Alert title
            message:
              type: string
              description: Alert message
    responses:
      200:
        description: Alert sent
    """
    try:
        data = request.get_json()
        alert_type = data.get('alert_type', 'manual')
        severity = data.get('severity', 'info')
        title = data.get('title', 'Manual Alert')
        message = data.get('message', '')

        result = alert_sender.send_alert(alert_type, severity, title, message)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Anomaly Detection Routes
# ============================================================

@app.route('/anomaly-detection')
@login_required
def anomaly_detection():
    """
    AI Anomaly Detection Dashboard
    ---
    tags:
      - Anomaly Detection
    responses:
      200:
        description: Anomaly detection dashboard page
    """
    stats = anomaly_detector.get_statistics()
    insights = anomaly_detector.get_insights()

    return render_template('anomaly-detection.html',
                         stats=stats,
                         insights=insights)

@app.route('/api/anomaly/stats')
@login_required
def anomaly_stats():
    """
    Get anomaly detection statistics
    ---
    tags:
      - Anomaly Detection
    responses:
      200:
        description: Anomaly statistics
        schema:
          type: object
          properties:
            success:
              type: boolean
            stats:
              type: object
    """
    try:
        stats = anomaly_detector.get_statistics()
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/anomaly/insights')
@login_required
def anomaly_insights():
    """
    Get AI insights from anomaly patterns
    ---
    tags:
      - Anomaly Detection
    responses:
      200:
        description: AI insights and recommendations
        schema:
          type: object
          properties:
            success:
              type: boolean
            insights:
              type: object
    """
    try:
        insights = anomaly_detector.get_insights()
        return jsonify({
            'success': True,
            'insights': insights
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/anomaly/list')
@login_required
def anomaly_list():
    """
    Get detected anomalies with optional filters
    ---
    tags:
      - Anomaly Detection
    parameters:
      - name: limit
        in: query
        type: integer
        default: 50
        description: Maximum number of anomalies to return
      - name: severity
        in: query
        type: string
        enum: [critical, high, medium, low]
        description: Filter by severity
      - name: resolved
        in: query
        type: boolean
        description: Filter by resolution status
    responses:
      200:
        description: List of anomalies
        schema:
          type: object
          properties:
            success:
              type: boolean
            anomalies:
              type: array
    """
    try:
        limit = request.args.get('limit', 50, type=int)
        severity = request.args.get('severity', None)
        resolved = request.args.get('resolved', None)

        # Convert resolved string to boolean
        if resolved is not None:
            resolved = resolved.lower() == 'true'

        anomalies = anomaly_detector.get_anomalies(
            limit=limit,
            severity=severity,
            resolved=resolved
        )

        return jsonify({
            'success': True,
            'anomalies': anomalies,
            'count': len(anomalies)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/anomaly/<anomaly_id>/acknowledge', methods=['POST'])
@login_required
def acknowledge_anomaly(anomaly_id):
    """
    Acknowledge an anomaly
    ---
    tags:
      - Anomaly Detection
    parameters:
      - name: anomaly_id
        in: path
        type: string
        required: true
        description: Anomaly ID
    responses:
      200:
        description: Anomaly acknowledged
      404:
        description: Anomaly not found
    """
    try:
        success = anomaly_detector.acknowledge_anomaly(anomaly_id)

        if success:
            return jsonify({
                'success': True,
                'message': 'Anomaly acknowledged successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Anomaly not found'
            }), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/anomaly/<anomaly_id>/resolve', methods=['POST'])
@login_required
def resolve_anomaly(anomaly_id):
    """
    Resolve an anomaly
    ---
    tags:
      - Anomaly Detection
    parameters:
      - name: anomaly_id
        in: path
        type: string
        required: true
        description: Anomaly ID
      - name: body
        in: body
        required: false
        schema:
          type: object
          properties:
            resolution_note:
              type: string
              description: Optional resolution note
    responses:
      200:
        description: Anomaly resolved
      404:
        description: Anomaly not found
    """
    try:
        data = request.get_json() or {}
        resolution_note = data.get('resolution_note', '')

        success = anomaly_detector.resolve_anomaly(anomaly_id, resolution_note)

        if success:
            return jsonify({
                'success': True,
                'message': 'Anomaly resolved successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Anomaly not found'
            }), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Predictive Analytics Routes
# ============================================================

@app.route('/predictive-analytics')
@login_required
def predictive_analytics_page():
    """
    Predictive Analytics Dashboard
    ---
    tags:
      - Predictive Analytics
    responses:
      200:
        description: Predictive analytics dashboard page
    """
    # Generate sample data for demonstration
    import random
    for i in range(100):
        predictive_analytics.add_metric_point('health_score', 70 + random.randint(-10, 20))
        predictive_analytics.add_metric_point('context_usage', 50 + random.randint(-20, 30))
        predictive_analytics.add_metric_point('error_count', random.randint(0, 15))
        predictive_analytics.add_metric_point('cost', 10 + random.uniform(-3, 5))
        predictive_analytics.add_metric_point('response_time', 200 + random.randint(-50, 100))

    return render_template('predictive-analytics.html')

@app.route('/api/forecast/summary')
@login_required
def forecast_summary():
    """
    Get forecast summary for all metrics
    ---
    tags:
      - Predictive Analytics
    responses:
      200:
        description: Forecast summary
        schema:
          type: object
          properties:
            success:
              type: boolean
            summary:
              type: object
    """
    try:
        summary = predictive_analytics.get_forecast_summary()
        return jsonify({
            'success': True,
            'summary': summary
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/forecast/insights')
@login_required
def forecast_insights():
    """
    Get predictive insights and recommendations
    ---
    tags:
      - Predictive Analytics
    responses:
      200:
        description: Predictive insights
        schema:
          type: object
          properties:
            success:
              type: boolean
            insights:
              type: object
    """
    try:
        insights = predictive_analytics.generate_forecast_insights()
        return jsonify({
            'success': True,
            'insights': insights
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/forecast/metric/<metric_name>')
@login_required
def forecast_metric(metric_name):
    """
    Get forecast for a specific metric
    ---
    tags:
      - Predictive Analytics
    parameters:
      - name: metric_name
        in: path
        type: string
        required: true
        description: Metric name to forecast
      - name: periods
        in: query
        type: integer
        default: 24
        description: Number of periods to forecast
      - name: method
        in: query
        type: string
        default: ensemble
        enum: [ensemble, linear, exponential, moving_average, seasonal]
        description: Forecasting method
    responses:
      200:
        description: Forecast data
        schema:
          type: object
    """
    try:
        periods = request.args.get('periods', 24, type=int)
        method = request.args.get('method', 'ensemble')

        forecast_result = predictive_analytics.forecast_metric(
            metric_name,
            periods=periods,
            method=method
        )

        return jsonify(forecast_result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/forecast/capacity-predictions')
@login_required
def capacity_predictions():
    """
    Get capacity breach predictions
    ---
    tags:
      - Predictive Analytics
    parameters:
      - name: horizon
        in: query
        type: integer
        default: 168
        description: Hours to look ahead (default 1 week)
    responses:
      200:
        description: Capacity predictions
        schema:
          type: object
          properties:
            success:
              type: boolean
            predictions:
              type: array
    """
    try:
        horizon = request.args.get('horizon', 168, type=int)

        predictions = []

        # Check capacity for key metrics
        thresholds = {
            'context_usage': 85,
            'error_count': 50,
            'cost': 100
        }

        for metric_name, threshold in thresholds.items():
            result = predictive_analytics.predict_capacity_breach(
                metric_name,
                threshold,
                horizon=horizon
            )

            if result['success'] and result.get('will_breach'):
                predictions.append(result)

        return jsonify({
            'success': True,
            'predictions': predictions,
            'horizon_hours': horizon
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/forecast/train-models', methods=['POST'])
@login_required
def train_forecast_models():
    """
    Train/retrain forecasting models
    ---
    tags:
      - Predictive Analytics
    responses:
      200:
        description: Models trained successfully
    """
    try:
        # Future: Implement model training
        return jsonify({
            'success': True,
            'message': 'Model training scheduled (placeholder for future ML models)'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Alert Routing and Escalation Routes
# ============================================================

@app.route('/alert-routing')
@login_required
def alert_routing_page():
    """
    Alert Routing Dashboard
    ---
    tags:
      - Alert Routing
    responses:
      200:
        description: Alert routing dashboard page
    """
    return render_template('alert-routing.html')

@app.route('/api/alert-routing/stats')
@login_required
def alert_routing_stats():
    """
    Get alert routing statistics
    ---
    tags:
      - Alert Routing
    responses:
      200:
        description: Alert routing statistics
    """
    try:
        stats = alert_routing.get_statistics()
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/active-alerts')
@login_required
def get_active_routed_alerts():
    """
    Get active alerts
    ---
    tags:
      - Alert Routing
    responses:
      200:
        description: List of active alerts
    """
    try:
        alerts = alert_routing.get_active_alerts()
        return jsonify({
            'success': True,
            'alerts': alerts
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/rules', methods=['GET', 'POST'])
@login_required
def routing_rules():
    """
    Get or create routing rules
    ---
    tags:
      - Alert Routing
    """
    if request.method == 'POST':
        try:
            data = request.get_json()
            rules_data = alert_routing.load_routing_rules()

            # Generate ID
            rule_id = f"rule_{len(rules_data['rules']) + 1}"
            data['id'] = rule_id

            rules_data['rules'].append(data)
            alert_routing.save_routing_rules(rules_data)

            return jsonify({
                'success': True,
                'message': 'Routing rule created',
                'rule_id': rule_id
            })
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    else:
        try:
            rules_data = alert_routing.load_routing_rules()
            return jsonify({
                'success': True,
                'rules': rules_data['rules']
            })
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/rules/<rule_id>/toggle', methods=['POST'])
@login_required
def toggle_routing_rule(rule_id):
    """Toggle routing rule enabled status"""
    try:
        rules_data = alert_routing.load_routing_rules()
        rule = next((r for r in rules_data['rules'] if r['id'] == rule_id), None)

        if not rule:
            return jsonify({'success': False, 'message': 'Rule not found'}), 404

        rule['enabled'] = not rule.get('enabled', True)
        alert_routing.save_routing_rules(rules_data)

        return jsonify({
            'success': True,
            'enabled': rule['enabled']
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/policies')
@login_required
def get_escalation_policies():
    """
    Get escalation policies
    ---
    tags:
      - Alert Routing
    """
    try:
        policies_data = alert_routing.load_escalation_policies()
        return jsonify({
            'success': True,
            'policies': policies_data['policies']
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/on-call-schedules')
@login_required
def get_on_call_schedules():
    """
    Get on-call schedules
    ---
    tags:
      - Alert Routing
    """
    try:
        schedules_data = alert_routing.load_on_call_schedules()

        # Get current on-call for each schedule
        current_on_call = {}
        for schedule in schedules_data['schedules']:
            current_on_call[schedule['id']] = alert_routing.get_current_on_call(schedule['id'])

        return jsonify({
            'success': True,
            'schedules': schedules_data['schedules'],
            'current_on_call': current_on_call
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/channels')
@login_required
def get_notification_channels():
    """
    Get notification channels
    ---
    tags:
      - Alert Routing
    """
    try:
        channels_data = alert_routing.load_notification_channels()
        return jsonify({
            'success': True,
            'channels': channels_data['channels']
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/alerts/<alert_id>/acknowledge', methods=['POST'])
@login_required
def acknowledge_routed_alert(alert_id):
    """Acknowledge a routed alert"""
    try:
        data = request.get_json() or {}
        acknowledged_by = data.get('acknowledged_by', 'admin')

        success = alert_routing.acknowledge_alert(alert_id, acknowledged_by)

        if success:
            return jsonify({
                'success': True,
                'message': 'Alert acknowledged'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Alert not found'
            }), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/alerts/<alert_id>/resolve', methods=['POST'])
@login_required
def resolve_routed_alert(alert_id):
    """Resolve a routed alert"""
    try:
        data = request.get_json() or {}
        resolved_by = data.get('resolved_by', 'admin')
        resolution_note = data.get('resolution_note', '')

        success = alert_routing.resolve_alert(alert_id, resolved_by, resolution_note)

        if success:
            return jsonify({
                'success': True,
                'message': 'Alert resolved'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Alert not found'
            }), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alert-routing/alerts/create', methods=['POST'])
@login_required
def create_routed_alert():
    """
    Create a new routed alert
    ---
    tags:
      - Alert Routing
    """
    try:
        data = request.get_json()
        alert = alert_routing.create_alert(data)

        return jsonify({
            'success': True,
            'message': 'Alert created and routed',
            'alert': alert
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================
# Memory System Integration Routes
# ============================================================

@app.route('/api/memory-system/health')
@login_required
def memory_system_health():
    """
    Get Claude Memory System v2.2.0 comprehensive health stats
    ---
    tags:
      - Memory System
    responses:
      200:
        description: Complete memory system health metrics
    """
    try:
        stats = memory_system_monitor.get_comprehensive_stats()
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/memory-system/daemons')
@login_required
def memory_system_daemons():
    """Get daemon status"""
    try:
        daemons = memory_system_monitor.get_daemon_status()
        return jsonify({
            'success': True,
            'daemons': daemons
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/memory-system/policies')
@login_required
def memory_system_policies():
    """Get policy status"""
    try:
        policies = memory_system_monitor.get_policy_status()
        return jsonify({
            'success': True,
            'policies': policies
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.errorhandler(404)
def page_not_found(e):
    """Handle 404 errors"""
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    """Handle 500 errors"""
    return render_template('500.html'), 500

@app.template_filter('format_datetime')
def format_datetime(value):
    """Format datetime for display"""
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except:
            return value

    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return value

@app.template_filter('format_number')
def format_number(value):
    """Format numbers with commas"""
    try:
        return '{:,}'.format(int(value))
    except:
        return value

# ============================================================
# WebSocket Event Handlers for Real-time Updates
# ============================================================

@socketio.on('connect')
def handle_connect():
    """Handle client connection"""
    print(f'Client connected: {request.sid}')
    emit('connection_response', {'status': 'connected', 'message': 'Connected to Claude Monitoring System'})

@socketio.on('disconnect')
def handle_disconnect():
    """Handle client disconnection"""
    print(f'Client disconnected: {request.sid}')

@socketio.on('request_metrics')
def handle_metrics_request():
    """Handle request for metrics data"""
    try:
        system_health = metrics.get_system_health()
        daemon_status = metrics.get_daemon_status()
        policy_status = policy_checker.get_detailed_policy_status()

        daemons_running = len([d for d in daemon_status if d.get('status') == 'running'])
        daemons_total = len(daemon_status) if daemon_status else 8
        health_score = system_health.get('health_score', system_health.get('score', 0))

        emit('metrics_update', {
            'health_score': health_score,
            'daemons_running': daemons_running,
            'daemons_total': daemons_total,
            'active_policies': policy_status.get('active_policies', 0),
            'context_usage': system_health.get('context_usage', 0),
            'memory_usage': system_health.get('memory_usage', 0),
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"Error emitting metrics: {e}")
        emit('error', {'message': str(e)})

# ============================================================
# WebSocket Handlers for Real-time Collaboration
# ============================================================

@socketio.on('collaboration:join')
def handle_collaboration_join(data):
    """Handle user joining a collaboration session"""
    try:
        session_id = data.get('session_id')
        widget_id = data.get('widget_id')
        username = session.get('username', 'admin')

        # Join session
        session_data = collaboration_manager.join_session(
            session_id=session_id,
            user_id=username,
            socket_id=request.sid
        )

        if session_data:
            # Notify other participants
            emit('collaboration:user_joined', {
                'session_id': session_id,
                'user_id': username,
                'color': next((p['color'] for p in session_data['participants'] if p['user_id'] == username), '#FF5733'),
                'timestamp': datetime.utcnow().isoformat() + 'Z'
            }, broadcast=True, include_self=False)

            # Send current session state to joiner
            emit('collaboration:session_state', {
                'session': session_data
            })
        else:
            emit('error', {'message': 'Session not found or expired'})

    except Exception as e:
        print(f"Error in collaboration:join: {e}")
        emit('error', {'message': str(e)})

@socketio.on('collaboration:leave')
def handle_collaboration_leave(data):
    """Handle user leaving a collaboration session"""
    try:
        session_id = data.get('session_id')
        username = session.get('username', 'admin')

        success = collaboration_manager.leave_session(
            session_id=session_id,
            user_id=username
        )

        if success:
            # Notify other participants
            emit('collaboration:user_left', {
                'session_id': session_id,
                'user_id': username,
                'timestamp': datetime.utcnow().isoformat() + 'Z'
            }, broadcast=True, include_self=False)

    except Exception as e:
        print(f"Error in collaboration:leave: {e}")
        emit('error', {'message': str(e)})

@socketio.on('collaboration:cursor_move')
def handle_cursor_move(data):
    """Handle cursor movement updates"""
    try:
        session_id = data.get('session_id')
        cursor_position = data.get('cursor_position')
        username = session.get('username', 'admin')

        # Update cursor position
        collaboration_manager.update_cursor(
            session_id=session_id,
            user_id=username,
            cursor_position=cursor_position
        )

        # Broadcast to other participants
        emit('collaboration:cursor_update', {
            'session_id': session_id,
            'user_id': username,
            'cursor_position': cursor_position,
            'timestamp': datetime.utcnow().isoformat() + 'Z'
        }, broadcast=True, include_self=False)

    except Exception as e:
        print(f"Error in collaboration:cursor_move: {e}")

@socketio.on('collaboration:edit')
def handle_collaboration_edit(data):
    """Handle edit operations"""
    try:
        session_id = data.get('session_id')
        operation = data.get('operation')
        username = session.get('username', 'admin')

        # Log operation
        collaboration_manager.log_operation(
            session_id=session_id,
            user_id=username,
            operation=operation
        )

        # Broadcast to other participants
        emit('collaboration:operation', {
            'session_id': session_id,
            'user_id': username,
            'operation': operation,
            'timestamp': datetime.utcnow().isoformat() + 'Z'
        }, broadcast=True, include_self=False)

    except Exception as e:
        print(f"Error in collaboration:edit: {e}")
        emit('error', {'message': str(e)})

@socketio.on('collaboration:lock_request')
def handle_lock_request(data):
    """Handle line lock requests"""
    try:
        session_id = data.get('session_id')
        editor = data.get('editor')
        line_range = tuple(data.get('line_range', [0, 0]))
        username = session.get('username', 'admin')

        # Request lock
        result = collaboration_manager.request_lock(
            session_id=session_id,
            user_id=username,
            editor=editor,
            line_range=line_range
        )

        if result.get('granted'):
            emit('collaboration:lock_granted', {
                'session_id': session_id,
                'lock_key': result.get('lock_key'),
                'editor': editor,
                'line_range': line_range
            })

            # Notify others
            emit('collaboration:lock_acquired', {
                'session_id': session_id,
                'user_id': username,
                'editor': editor,
                'line_range': line_range
            }, broadcast=True, include_self=False)
        else:
            emit('collaboration:conflict', {
                'message': result.get('reason'),
                'locked_by': result.get('locked_by')
            })

    except Exception as e:
        print(f"Error in collaboration:lock_request: {e}")
        emit('error', {'message': str(e)})

@socketio.on('collaboration:chat')
def handle_collaboration_chat(data):
    """Handle chat messages in collaboration session"""
    try:
        session_id = data.get('session_id')
        message = data.get('message')
        username = session.get('username', 'admin')

        # Broadcast chat message
        emit('collaboration:chat_message', {
            'session_id': session_id,
            'user_id': username,
            'message': message,
            'timestamp': datetime.utcnow().isoformat() + 'Z'
        }, broadcast=True)

    except Exception as e:
        print(f"Error in collaboration:chat: {e}")
        emit('error', {'message': str(e)})

# Background thread for real-time updates
def background_thread():
    """Background thread to emit real-time updates every 10 seconds"""
    while True:
        time.sleep(10)  # Update every 10 seconds
        try:
            system_health = metrics.get_system_health()
            daemon_status = metrics.get_daemon_status()
            policy_status = policy_checker.get_detailed_policy_status()

            daemons_running = len([d for d in daemon_status if d.get('status') == 'running'])
            daemons_total = len(daemon_status) if daemon_status else 8
            health_score = system_health.get('health_score', system_health.get('score', 0))

            socketio.emit('metrics_update', {
                'health_score': health_score,
                'daemons_running': daemons_running,
                'daemons_total': daemons_total,
                'active_policies': policy_status.get('active_policies', 0),
                'context_usage': system_health.get('context_usage', 0),
                'memory_usage': system_health.get('memory_usage', 0),
                'timestamp': datetime.now().isoformat()
            }, broadcast=True)
        except Exception as e:
            print(f"Error in background thread: {e}")

# Start background thread
thread = threading.Thread(target=background_thread, daemon=True)
thread.start()

if __name__ == '__main__':
    print("""
    ============================================================
    Claude Monitoring System v2.12 (Memory Integration Edition)
    ============================================================

    Dashboard URL: http://localhost:5000
    API Docs: http://localhost:5000/api/docs
    Widget Builder: http://localhost:5000/widget-builder
    Community: http://localhost:5000/community-marketplace
    AI Detection: http://localhost:5000/anomaly-detection
    Forecasting: http://localhost:5000/predictive-analytics
    Alert Routing: http://localhost:5000/alert-routing
    Username: admin
    Password: admin

    🧠 Memory System v2.2.0 Integration:
    ✓ 8 Daemon health monitoring (context, auto-save, git, etc.)
    ✓ 10 Policy enforcement tracking (active status)
    ✓ Context optimization metrics (cache hits, token savings)
    ✓ Failure prevention stats (auto-fixes, patterns learned)
    ✓ Model selection distribution (haiku/sonnet/opus usage)
    ✓ Session memory tracking (active sessions, pruning)
    ✓ Git auto-commit activity monitoring
    ✓ Overall system health score calculation

    🚀 Advanced Features:
    ✓ Custom alert routing & escalation policies
    ✓ Multi-level escalation chains (up to 3 levels)
    ✓ Predictive analytics & forecasting (5 algorithms)
    ✓ AI-powered anomaly detection (6 ML algorithms)
    ✓ Community widget marketplace with sharing
    ✓ Email & SMS alerts for critical issues
    ✓ Custom dashboard themes (6 themes)
    ✓ Real-time WebSocket updates (10s interval)

    Starting server...
    ============================================================
    """)

    socketio.run(app, debug=True, host='0.0.0.0', port=5000, allow_unsafe_werkzeug=True)
